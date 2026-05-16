import json
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.config import settings
from app.schemas.pipeline import ProcessedRecord

HEADERS = [
    "ID", "Source", "Timestamp", "Model",
    "Category", "Priority", "Confidence",
    "Core Issue", "Identifiers", "Urgency Signal",
    "Destination Queue", "Escalated", "Escalation Reason",
    "Summary",
]

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ESCALATED_FILL = PatternFill("solid", fgColor="FFF3CD")
COL_WIDTHS = [14, 14, 22, 20, 18, 10, 12, 40, 36, 36, 18, 10, 36, 60]


def _output_path() -> Path:
    path = Path(settings.output_dir) / "processed_records.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _record_to_row(r: ProcessedRecord) -> list:
    return [
        r.id,
        r.source,
        r.timestamp,
        r.model_used,
        r.classification.category,
        r.classification.priority,
        round(r.classification.confidence_score, 2),
        r.enrichment.core_issue,
        json.dumps(r.enrichment.identifiers),
        r.enrichment.urgency_signal,
        r.routing.destination_queue,
        "Yes" if r.routing.escalation_flag else "No",
        r.routing.escalation_reason or "",
        r.summary,
    ]


def _apply_header(ws) -> None:
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def upsert_record(record: ProcessedRecord) -> None:
    path = _output_path()
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Processed Records"
        _apply_header(ws)

    row_data = _record_to_row(record)
    row_idx = ws.max_row + 1
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    if record.routing.escalation_flag:
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col).fill = ESCALATED_FILL

    ws.row_dimensions[row_idx].height = 48
    wb.save(path)
